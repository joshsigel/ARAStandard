#!/usr/bin/env python3
"""
ARA Standard v1.1 — Regulatory Mappings Seeding Script
Reads the REG Master Excel and seeds the regulatory_mappings table in Supabase.

Data sources:
  1. Per-framework sheets (14 frameworks) — ACR-to-framework_requirement_id mappings
  2. Master Crosswalk sheet — platform_cert_sufficient values (Y/N/Partial)

Usage:
  python3 scripts/seed-regulatory-mappings.py

Requires:
  - SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY in .env.local
  - openpyxl package (pip install openpyxl)
"""

import json
import os
import re
import sys
from collections import Counter

# ─── Load .env.local ─────────────────────────────────────────────────────────

env_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '.env.local')
if os.path.exists(env_path):
    with open(env_path) as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                key, value = line.split('=', 1)
                os.environ[key.strip()] = value.strip()

try:
    from supabase import create_client, Client
except ImportError:
    print("Installing supabase-py...")
    os.system(f"{sys.executable} -m pip install supabase")
    from supabase import create_client, Client

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl

SUPABASE_URL = os.environ.get('NEXT_PUBLIC_SUPABASE_URL', '')
SUPABASE_KEY = os.environ.get('SUPABASE_SERVICE_ROLE_KEY', '')

if not SUPABASE_URL or not SUPABASE_KEY:
    print("ERROR: Set NEXT_PUBLIC_SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY in .env.local")
    sys.exit(1)

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# ─── Configuration ────────────────────────────────────────────────────────────

EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    'Documents', '002', 'ARAF_ARA_REG_2026_002_Master.xlsx'
)

# Map from per-framework sheet name to database framework_id.
# Column 4 in each per-framework sheet contains the framework requirement reference.
FRAMEWORK_SHEETS = {
    'NIST AI RMF':   'nist-ai-rmf',
    'EU AI Act':     'eu-ai-act',
    'ISO 42001':     'iso-42001',
    'ISO 27001':     'iso-27001',
    'NIST 800-53':   'nist-800-53',
    'SOC 2':         'soc-2',
    'IEC 61508':     'iec-61508',
    'IEC 62443':     'iec-62443',
    'OWASP LLM':    'owasp-llm',
    'MITRE ATLAS':   'mitre-atlas',
    'CSA AICM':      'csa-aicm',
    'ISO 42005':     'iso-42005',
    'ISO 42006':     'iso-42006',
    'CoSAI':         'cosai',
}


# ─── Extract mappings from Excel ─────────────────────────────────────────────

def extract_mappings():
    """
    Extract all regulatory mappings from the Excel workbook.

    The workbook structure:
      - Master Crosswalk: ACR ID (col 1), framework refs (cols 7-15),
        platform_cert_sufficient (col 16)
      - Per-framework sheets: ACR ID (col 1), framework requirement ref (col 4)
        Each row is one ACR-to-requirement mapping. ACRs with multiple framework
        requirements appear on separate rows.

    Returns a list of dicts ready for insertion.
    """
    print(f"Loading workbook: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # Step 1: Build platform_cert_sufficient lookup from Master Crosswalk
    print("\nReading Master Crosswalk for platform_cert_sufficient values...")
    ws = wb['Master Crosswalk']
    cert_sufficient = {}  # acr_id -> 'Y' | 'N' | 'Partial'
    for row_idx in range(5, ws.max_row + 1):  # Data starts at row 5 (header at row 4)
        acr_id = ws.cell(row=row_idx, column=1).value
        pcs = ws.cell(row=row_idx, column=16).value
        if acr_id and str(acr_id).startswith('ACR-') and pcs:
            cert_sufficient[acr_id] = str(pcs).strip()
    print(f"  Found {len(cert_sufficient)} ACRs with platform_cert_sufficient values")

    # Step 2: Extract mappings from each per-framework sheet
    all_mappings = []

    for sheet_name, framework_id in FRAMEWORK_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            print(f"\n  WARNING: Sheet '{sheet_name}' not found, skipping")
            continue

        ws = wb[sheet_name]
        sheet_mappings = []

        # Scan all rows for ACR entries (col 1 = ACR-X.XX, col 4 = framework ref)
        for row_idx in range(1, ws.max_row + 1):
            acr_id = ws.cell(row=row_idx, column=1).value
            fw_ref = ws.cell(row=row_idx, column=4).value

            if not acr_id or not str(acr_id).startswith('ACR-'):
                continue
            if not fw_ref or str(fw_ref).strip() in ('—', '-', '', 'None'):
                continue

            acr_id = str(acr_id).strip()
            fw_ref = str(fw_ref).strip()

            # Get platform_cert_sufficient for this ACR
            pcs = cert_sufficient.get(acr_id, 'N')

            sheet_mappings.append({
                'acr_id': acr_id,
                'framework_id': framework_id,
                'framework_requirement_id': fw_ref,
                'mapping_type': 'direct',
                'platform_cert_sufficient': pcs,
            })

        print(f"  {sheet_name:20s} ({framework_id:15s}): {len(sheet_mappings):4d} mappings")
        all_mappings.extend(sheet_mappings)

    wb.close()
    return all_mappings


# ─── Validate mappings ───────────────────────────────────────────────────────

def validate_mappings(mappings):
    """Validate ACR IDs and framework IDs before insertion."""
    print("\nValidating mappings...")

    acr_ids = sorted(set(m['acr_id'] for m in mappings))
    print(f"  Unique ACR IDs referenced: {len(acr_ids)}")
    print(f"  ACR range: {acr_ids[0]} to {acr_ids[-1]}")

    fw_ids = sorted(set(m['framework_id'] for m in mappings))
    print(f"  Unique framework IDs: {len(fw_ids)}")
    print(f"  Frameworks: {', '.join(fw_ids)}")

    # Check for exact duplicates
    seen = set()
    dupes = 0
    for m in mappings:
        key = (m['acr_id'], m['framework_id'], m['framework_requirement_id'])
        if key in seen:
            dupes += 1
        seen.add(key)
    if dupes > 0:
        print(f"  WARNING: {dupes} exact duplicate rows (will be deduplicated)")
    else:
        print(f"  No duplicates found")

    return True


# ─── Seed to Supabase ────────────────────────────────────────────────────────

def seed_mappings(mappings):
    """Insert mappings into the regulatory_mappings table."""
    print("\n" + "=" * 60)
    print("Seeding regulatory_mappings table")
    print("=" * 60)

    # Deduplicate by (acr_id, framework_id, framework_requirement_id)
    seen = set()
    unique_mappings = []
    for m in mappings:
        key = (m['acr_id'], m['framework_id'], m['framework_requirement_id'])
        if key not in seen:
            seen.add(key)
            unique_mappings.append(m)

    print(f"\nTotal unique mappings to insert: {len(unique_mappings)}")
    if len(unique_mappings) != len(mappings):
        print(f"  (Deduplicated from {len(mappings)} raw rows)")

    # Clear existing data
    print("\nClearing existing regulatory_mappings data...")
    try:
        result = supabase.table('regulatory_mappings').delete().neq('id', 0).execute()
        print(f"  Cleared existing rows")
    except Exception as e:
        print(f"  Note: {e}")

    # Insert in batches of 500
    BATCH_SIZE = 500
    total_inserted = 0

    for i in range(0, len(unique_mappings), BATCH_SIZE):
        batch = unique_mappings[i:i + BATCH_SIZE]
        try:
            result = supabase.table('regulatory_mappings').insert(batch).execute()
            total_inserted += len(result.data)
            print(f"  Inserted batch {i // BATCH_SIZE + 1}: {len(result.data)} rows (total: {total_inserted})")
        except Exception as e:
            print(f"  ERROR inserting batch {i // BATCH_SIZE + 1}: {e}")
            # Fall back to row-by-row insertion
            for j, row in enumerate(batch):
                try:
                    result = supabase.table('regulatory_mappings').insert(row).execute()
                    total_inserted += 1
                except Exception as e2:
                    print(f"    FAILED: {row['acr_id']} / {row['framework_id']} / "
                          f"{row['framework_requirement_id']}: {e2}")

    print(f"\nTotal inserted: {total_inserted}")
    return total_inserted


# ─── Verify ──────────────────────────────────────────────────────────────────

def verify():
    """Verify row counts per framework using exact count queries."""
    print("\n" + "=" * 60)
    print("Verification — Row counts per framework")
    print("=" * 60)

    # Total count
    result = supabase.table('regulatory_mappings').select('id', count='exact').execute()
    print(f"\nTotal rows: {result.count}")

    # Count per framework (using count='exact' to avoid default row limit)
    frameworks = sorted(FRAMEWORK_SHEETS.values())
    total = 0
    print("\nMappings per framework:")
    for fw_id in frameworks:
        result = supabase.table('regulatory_mappings').select('id', count='exact').eq('framework_id', fw_id).execute()
        count = result.count
        print(f"  {fw_id:20s}: {count:4d}")
        total += count
    print(f"  {'':20s}  ----")
    print(f"  {'TOTAL':20s}: {total:4d}")

    # Platform cert sufficient distribution
    print("\nPlatform cert sufficient distribution:")
    for pcs in ['Y', 'N', 'Partial']:
        result = supabase.table('regulatory_mappings').select('id', count='exact').eq('platform_cert_sufficient', pcs).execute()
        print(f"  {pcs:10s}: {result.count:4d}")


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("ARA Standard — Regulatory Mappings Seeder")
    print("=" * 60)

    mappings = extract_mappings()
    print(f"\nTotal raw mappings extracted: {len(mappings)}")

    validate_mappings(mappings)
    seed_mappings(mappings)
    verify()

    print("\n" + "=" * 60)
    print("Seeding complete!")
    print("=" * 60)


if __name__ == '__main__':
    main()
